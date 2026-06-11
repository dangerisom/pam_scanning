"""Core PAM-scanning algorithm library.

This module contains the validated computational method behind the PAM-scanning
manuscript: sequence utilities, codon-table handling, PAM-site discovery,
PAM- and guide-silencing, BLAST+ off-target evaluation, optimal-guide selection,
scannable-sequence calculation, and primer-order generation.

The scientific logic here is intentionally preserved verbatim; only dead code has
been removed and the previously missing ``biomek`` plate dependency has been
replaced by the vendored :mod:`pam_scanning.plates` module. Progress is reported
via ``print`` so it appears on stdout when driven from the CLI or GUI.
"""

singleLetterAA = {"ALA":"A", "ASP":"D", "ASN":"N", "ARG":"R", "CYS":"C", "GLY":"G", "GLU":"E", "GLN":"Q", "HIS":"H", "ILE":"I", "LEU":"L", "LYS":"K", "MET":"M", "PRO":"P", "PHE":"F", "SER":"S", "THR":"T", "TYR":"Y", "TRP":"W", "VAL":"V"}

def reverseComplement(sequence):

	# Calculate the reverse complement sequence
	sequenceComplement = []
	i = 0
	for b in sequence:
		if b == "G":
			sequenceComplement.append("C")
		elif b == "C":
			sequenceComplement.append("G")
		elif b == "A":
			sequenceComplement.append("T")
		elif b == "T":
			sequenceComplement.append("A")
		elif b == "g":
			sequenceComplement.append("c")
		elif b == "c":
			sequenceComplement.append("g")
		elif b == "a":
			sequenceComplement.append("t")
		elif b == "t":
			sequenceComplement.append("a")
	sequenceComplement.reverse()
	sequenceReverseComplement = ""
	for b in sequenceComplement:
		sequenceReverseComplement += b
	return sequenceReverseComplement

def complement(sequence):

	# Calculate the reverse complement sequence
	sequenceComplement = []
	for b in sequence:
		if b == "G":
			sequenceComplement.append("C")
		elif b == "C":
			sequenceComplement.append("G")
		elif b == "A":
			sequenceComplement.append("T")
		elif b == "T":
			sequenceComplement.append("A")
		elif b == "g":
			sequenceComplement.append("c")
		elif b == "c":
			sequenceComplement.append("g")
		elif b == "a":
			sequenceComplement.append("t")
		elif b == "t":
			sequenceComplement.append("a")
	sequenceComplementString = ""
	for b in sequenceComplement:
		sequenceComplementString += b
	return sequenceComplementString

def fasta(sequence):

	i, formattedSequence = 0, ""
	while i < len(sequence):
		j = 1
		while j <= 60:
			formattedSequence += sequence[i]
			j += 1
			i += 1
			if i >= len(sequence):
				break
		formattedSequence += "\n"
	return formattedSequence

def countMismatches(sequence1, sequence2):
	i, mismatches = 0, 0
	while i < len(sequence1):
		if sequence1[i] != sequence2[i]:
			mismatches += 1
		i += 1
	return mismatches

def markSilencers(sequence1, sequence2):

	# sequence1 is the updated sequence...
	# sequence2 is the original sequence...

	# If the sequences are different lengths it is because one of the sequences contains an intentional mutation (at its 5' end).
	# Add leading spaces to the un-mutated sequence so that both sequences are the same length. 
	if len(sequence1) < len(sequence2):
		sequence1 += (len(sequence2)-len(sequence1))*" "
	elif len(sequence2) < len(sequence1):
		sequence2 += (len(sequence1)-len(sequence2))*" "

	i, markedSequence = 0, ""
	while i < len(sequence1):
		# Skip any leading black spaces added to sequence1 due to introduced mutations...
		if sequence1[i] == " ":
			i += 1
			continue
		# Skip any leading black spaces added to sequence2 due to introduced mutations...
		if sequence2[i] == " ":
			i += 1
			continue
		if sequence1[i] == sequence2[i]:
			markedSequence += sequence1[i].lower()
		else:
			markedSequence += sequence1[i]
		i += 1
	return markedSequence

##########################################################################################################################################################
# Parse the yeast codon table. Source: X
##########################################################################################################################################################
byRes, byCodon = {}, {}
def set_codon_table(codon_table_file_path):

	inputFile = open(codon_table_file_path, "r") 
	fileLines = inputFile.readlines()
	# byRes, byCodon = {}, {}
	for fileLine in fileLines[10:]:
		cData = fileLine.split()
		res, codon, percent = cData[0].upper(), cData[1], round(float(cData[2])*100, 1)
		byCodon[codon] = res
		if res not in byRes:
			byRes[res] = {(percent,codon):codon} # XXXX Changed this and it may mess things up...
		else:
			byRes[res].update({(percent,codon):codon})

# def blastGuides(outputPath, geneName, guides, localBlastDb, localGenomeFilePath, pamInclusionSequenceThreshold=6, pamInclusionThreshold=5, final=0):
def blastGuides(outputPath, geneName, guides, localBlastDb, localGenomeFilePath, pamInclusionSequenceThreshold=15, pamInclusionThreshold=5, final=0):

	if final:
		print("BLAST+: Calculating off-target potential for each FINAL guide sequence...")
	else:
		print("BLAST+: Calculating off-target potential for each guide sequence...")

	# Arguments...
	#
	# outputPath: X
	# geneName = X
	# guides = X
	# localBlastDb = X
	# pamInclusionSequenceThreshold = X
	# pamInclusionThreshold = Number of allowable PAM inclusions per guide. 

	# This function will only work if you have NCBI BLAST+ 
	# and the reference genome(s) installed on your computer...
	# For instructions on how to set these things up, see blast_README.txt...
	#
	# Current localBlastDb options...
	# "yeast" = S. cerevisiae (details)
	# "pombe" = S. pombe (details)
	# "human" = H. sapien (details)

	########################################################################################
	# Open and parse the reference genome file...
	########################################################################################	

	genomeFile = open(localGenomeFilePath)
	lines = genomeFile.readlines()
	i, genome = 0, {}
	while i < len(lines) - 1:
		line = lines[i]
		if line[0:4] == ">chr":
			chromosome = int(line.split(">chr")[1].split()[0])
			sequence = lines[i+1].strip()
			genome[chromosome] = sequence
		i += 1

	########################################################################################
	# Get the guides...
	########################################################################################

	guideKeys = sorted(guides)
	if final:
		guideFasta = open(outputPath + geneName + "-guidesBlast-final.fa", "w")
	else:
		guideFasta = open(outputPath + geneName + "-guidesBlast.fa", "w")
	for guideKey in guideKeys:
		guideFasta.write("> " + str(guideKey[0]) + " " + guides[guideKey] + " guide query" + "\n")
		guideFasta.write(guides[guideKey] + "\n")
	guideFasta.close()

	########################################################################################
	# BLAST the guides against a locally built genome database...
	########################################################################################

	from subprocess import run
	if final:
		run(["blastn", "-query", outputPath + geneName + "-guidesBlast-final.fa", "-task", "blastn-short", "-db", localBlastDb, "-out", outputPath + geneName + "-guidesBlastResult-final.txt", "-outfmt", "1"])
	else:
		run(["blastn", "-query", outputPath + geneName + "-guidesBlast.fa", "-task", "blastn-short", "-db", localBlastDb, "-out", outputPath + geneName + "-guidesBlastResult.txt", "-outfmt", "1"])

	########################################################################################
	# Parse results for threats of off-targeting cutting...
	########################################################################################

	# Open and Parse the BLAST results file...
	if final:
		blastResults = open(outputPath + geneName + "-guidesBlastResult-final.txt", "r")
	else:
		blastResults = open(outputPath + geneName + "-guidesBlastResult.txt", "r")
	safeGuides, unsafeGuides, keyedGuides = {}, {}, {}
	i, collect, safeGuide, guideSequence, guideStartIndex, guideStopIndex, lines, pamInclusions, pamInclusionsDict, pamInclusionsDictTmp = 0, 0, 1, None, 0, 0, blastResults.readlines(), 0, {}, {}
	superConservativePamInclusionDict, superConservativePamInclusionDictTmp = {}, {}
	partialGuide, internalGuideStartIndex, internalGuideStopIndex = None, 0, 0
	while i < len(lines):

		# Get the file line...
		line = lines[i]
		sLine = line.split()

		# Get the guide key and guide sequence needed for the return value...
		if "Query=" in line:
			guideKey = int(line.split()[1])
			guideSequence = line.split()[2]
			print("BLAST+: Evaluating off-target BLAST results for guide sequence: ", guideSequence)

		# Occasionally, no BLAST results are returned for a query sequence...
		# In such a case, the guide is 100% safe...
		if "***** No hits found *****" in line:
			safeGuide = 1
			i += 1
			continue

		# Conditional trigger to evaluate whether a guide sequence is safe...
		# This condition is triggered once all of the BLAST hits for a given guide sequence have been evaluated...
		if not sLine and collect:

			# Enforce PAM inclusion threshold...
			if pamInclusions:
				# Save a record for monitoring CRISPR on- and off-target performance...
				pamInclusionsDict.update(pamInclusionsDictTmp)
				if pamInclusions > pamInclusionThreshold:
					safeGuide = 0

			# Evaluate guide safety and save safe guide sequences...
			if safeGuide:
				safeGuides[(guideKey, guideSequence)] = guideSequence
			else:
				unsafeGuides[(guideKey, guideSequence)] = guideSequence

			# Collect super conservative PAM inclusions...
			superConservativePamInclusionDict.update(superConservativePamInclusionDictTmp)

			# Reset loop collection variables...
			safeGuide = 1
			collect = 0
			guideSequence = None
			guideStartIndex = 0
			guideStopIndex = 0
			pamInclusions = 0
			pamInclusionsDictTmp = {}
			superConservativePamInclusionDictTmp = {}
			partialGuide, internalGuideStartIndex, internalGuideStopIndex = None, 0, 0

		# File line condition for a new guide BLAST query result for a specific guide sequence...
		if "Query_" in line:

			# Collect the BLAST hit results for a given guide sequence query... 
			collect = 1
			line = line.upper() # In case lower case bases are present in the localBlastDb file...

			# For heterologous genes that don't necessarily have a self-match in the reference genome...
			if guideSequence not in line:

				# Guide query may be a sub-sequence of original guide sequence...
				partialGuide = line.split()[2]

				# XXXX

				# If necessary, adjust the partial guide to have leading and trailing blank characters...
				# This ensures the partial guide sequence and original guide sequence are the same length...
				# This also aligns the matching portion of the partial guide within the original guide sequence...
				internalGuideStartIndex = guideSequence.index(partialGuide)
				internalGuideStopIndex = internalGuideStartIndex + len(partialGuide)
				# tmpGuide = partialGuide
				# if internalGuideStartIndex > 0:
				# 	tmpGuide = internalGuideStartIndex*" " + tmpGuide  
				# if internalGuideStopIndex <= len(guideSequence) - 1:
				# 	tmpGuide = tmpGuide + (len(guideSequence) - internalGuideStopIndex)*" "
				# guideStartIndex = line.index(tmpGuide)
				# guideStopIndex = guideStartIndex + len(tmpGuide)

				# XXXX

				guideStartIndex = line.index(partialGuide)
				guideStopIndex = guideStartIndex + len(partialGuide)
				i += 1 # Increment by 1 because there is no self-match if a heterologous gene integrated in the reference genome...
			
			# For guide sequences within genes that will have a self-match in the reference genome...
			else:

				# In the case of a self-match, the guide query sequence is ALWAYS the same as the original guide sequence...
				guideStartIndex = line.index(guideSequence)
				guideStopIndex = guideStartIndex + len(guideSequence)
				i += 2 # Increment by 2 because the first result in the BLAST file is the self-match...
			continue

		# Evaluate each individual BLAST hit for the specific guide sequence...
		if collect:

			# In case of insertion, continue...
			# For example:
			#
			# 0        67734    ....................     67754
			#                           \              
			#                           |              
			#                           G
			# 0        28283     ............            28294
			#
			# In this case, the \, |, and G lines are skipped...
			####################################################
			if line[0] == " ":
				i += 1
				continue

			# Get the matching portion of the BLAST result for the guide sequence...
			match = line[guideStartIndex:guideStopIndex]
			# XXXX
			if partialGuide:
				tmpGuide = match #line.split()[2] #partialGuide
				if internalGuideStartIndex > 0:
					tmpGuide = internalGuideStartIndex*" " + tmpGuide  
				if internalGuideStopIndex <= len(guideSequence) - 1:
					tmpGuide = tmpGuide + (len(guideSequence) - internalGuideStopIndex)*" "
				match = tmpGuide

				# print("*", guideSequence, "*")
				# print("*", match, "*")
				# print(line)

			# XXXX

			c, blastHit, blastHitRaw, nMatches, pseudoMatchingSequence = 0, "", "", 0, ""
			while c < len(guideSequence):
				mc = match[c]
				if mc != " ":
					# Pseudomatch because the BLAST hit may contain base letter mismatches...
					# Maintaining these mismatches, which are included in the mismatch count,...
					# ...preserves the substring search functionality necessary for sub-guide positioning...
					pseudoMatchingSequence += guideSequence[c]
				if mc == " ":
					blastHit += match[c] # include blank spaces...
				elif mc != ".":
					blastHit += match[c] # include similar base mismatches...
					blastHitRaw += match[c]
				elif mc == ".":
					blastHit += guideSequence[c] # replace . with the sequence bases...
					blastHitRaw += guideSequence[c]
					nMatches += 1
				c += 1

			# print("*", guideSequence, "*")
			# print("*", blastHit, "*")
			# print(line)

			# XXXX

			# If more than 17 bases match between the BLAST hit and the guide sequence... 
			if nMatches > 17:

				# Check if the BLAST hit even has a PAM site...
				# If it does NOT have a PAM, it is a safe guide sequence...
				# If it DOES have a PAM, check from PAM inclusions...

				if guideSequence[0:2] == "CC":

					# # Do not include the first two guide bases in the calculation...
					# n, matchingSequence = 0, blastHit[0:2]
					# while n < len(guideSequence[2:]):
					# 	if guideSequence[2:][n] == blastHit[2:][n]:
					# 		matchingSequence += guideSequence[2:][n]
					# 	else:
					# 		break
					# 	n += 1

					n, matchingSequence = 0, ""
					while n < len(guideSequence) - 1:
						if guideSequence[n] == blastHit[n]:
							matchingSequence += guideSequence[n]
						else:
							break
						n += 1  

					# Guide match with PAM...
					# The only BLAST hits to worry about are those that include a PAM (or wobble PAM)...
					if blastHit[0:2] in ["CC", "CT", "TC", " C", " T"]: # Catching PAM and PAM wobble...
					#if blastHit[0:2] in ["CC", "CT", "TC"]: # Catching PAM and PAM wobble...

						# Have to extend into the genome to check PAM wobble...
						if blastHit[0] == " ":

							blastCheck = blastHitRaw
							gKeys = sorted(genome)
							for gKey in gKeys:
								result = genome[gKey].find(blastCheck)
								if result != -1:
									resultIndex = genome[gKey].index(blastCheck)
									blastCheckExtended = genome[gKey][resultIndex-1:resultIndex+len(blastCheck)]
									# print()
									# print(1, blastCheck)
									# print(2, blastCheckExtended)
									if blastCheckExtended[0] in "CT":

										# Conservative...
										if len(matchingSequence) > pamInclusionSequenceThreshold:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											pamInclusionsDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
											pamInclusions += 1
										# Super conservative...
										if len(matchingSequence) > 9:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											superConservativePamInclusionDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)

								result = genome[gKey].find(reverseComplement(blastCheck))
								if result != -1:
									resultIndex = genome[gKey].index(reverseComplement(blastCheck))
									blastCheckExtended = genome[gKey][resultIndex:resultIndex+len(blastCheck)+1]
									# print()
									# print("reversed...")
									# print(1, reverseComplement(blastCheck))
									# print(2, blastCheckExtended)
									if blastCheckExtended[-1] in "AG":

										# Conservative...
										if len(matchingSequence) > pamInclusionSequenceThreshold:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											pamInclusionsDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
											pamInclusions += 1
										# Super conservative...
										if len(matchingSequence) > 9:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											superConservativePamInclusionDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)

						else:

							# Conservative...
							if len(matchingSequence) > pamInclusionSequenceThreshold:
								# For monitoring PAM inclusions as we learn more about CRISPR...
								pamInclusionsDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
								pamInclusions += 1
							# Super conservative...
							if len(matchingSequence) > 9:
								# For monitoring PAM inclusions as we learn more about CRISPR...
								superConservativePamInclusionDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)

					# BLAST hits without PAMs, which are safe...
					# Nothing needs to be done...
					else:
						pass

				else:

					# # Do not include the last two guide bases in the calculation...
					# n, matchingSequence = len(guideSequence) - 3, blastHit[-2:]
					# while n >= 0:
					# 	if guideSequence[0:-2][n] == blastHit[0:-2][n]:
					# 		matchingSequence += guideSequence[0:-2][n]
					# 	else:
					# 		break
					# 	n -= 1

					n, matchingSequence = len(guideSequence) - 1, ""
					while n >= 0:
						if guideSequence[n] == blastHit[n]:
							matchingSequence += guideSequence[n]
						else:
							break
						n -= 1

					# Guide match with PAM...
					# The only BLAST hits to worry about are those that include a PAM (or wobble PAM)...
					if blastHit[-2:] in ["GG", "GA", "AG", "G ", "A "]: # Catching PAM and PAM wobble...
					#if blastHit[-2:] in ["GG", "GA", "AG"]: # Catching PAM and PAM wobble...

						# Have to extend into the genome to check PAM wobble...
						if blastHit[-1] == " ":

							blastCheck = blastHitRaw
							gKeys = sorted(genome)
							for gKey in gKeys:
								result = genome[gKey].find(blastCheck)
								if result != -1:
									resultIndex = genome[gKey].index(blastCheck)
									#blastCheckExtended = genome[gKey][resultIndex-1:resultIndex+len(blastCheck)]
									blastCheckExtended = genome[gKey][resultIndex:resultIndex+len(blastCheck)+1]
									# print()
									# print(1, blastCheck)
									# print(2, blastCheckExtended)
									if blastCheckExtended[0] in "AG":

										# Conservative...
										if len(matchingSequence) > pamInclusionSequenceThreshold:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											pamInclusionsDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
											pamInclusions += 1
										# Super conservative...
										if len(matchingSequence) > 9:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											superConservativePamInclusionDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)

								result = genome[gKey].find(reverseComplement(blastCheck))
								if result != -1:
									resultIndex = genome[gKey].index(reverseComplement(blastCheck))
									#blastCheckExtended = genome[gKey][resultIndex:resultIndex+len(blastCheck)+1]
									blastCheckExtended = genome[gKey][resultIndex-1:resultIndex+len(blastCheck)]
									# print()
									# print("reversed...2")
									# print(1, reverseComplement(blastCheck))
									# print(2, blastCheckExtended)
									if blastCheckExtended[-1] in "CT":

										# Conservative...
										if len(matchingSequence) > pamInclusionSequenceThreshold:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											pamInclusionsDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
											pamInclusions += 1
										# Super conservative...
										if len(matchingSequence) > 9:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											superConservativePamInclusionDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)

						else:

							# Conservative...
							if len(matchingSequence) > pamInclusionSequenceThreshold:
								# For monitoring PAM inclusions as we learn more about CRISPR...
								pamInclusionsDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
								pamInclusions += 1
							# Super conservative...
							if len(matchingSequence) > 9:
								# For monitoring PAM inclusions as we learn more about CRISPR...
								superConservativePamInclusionDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)

					# BLAST hits without PAMs, which are safe...
					# Nothing needs to be done...
					else:
						pass

			# Some BLAST hits are shorter than 17 bases, so they can never meet the threshold of > 17 nMatches...
			# However, these short BLAST hits could qualify as a PAM inclusion...
			# As such, check short BLAST hits for PAM inclusions...			
			else:

				if guideSequence[0:2] == "CC":

					# # Do not include the first two guide bases in the calculation...
					# n, matchingSequence = 0, blastHit[0:2]
					# while n < len(guideSequence[2:]):
					# 	if guideSequence[2:][n] == blastHit[2:][n]:
					# 		matchingSequence += guideSequence[2:][n]
					# 	else:
					# 		break
					# 	n += 1

					n, matchingSequence = 0, ""
					while n < len(guideSequence) - 1:
						if guideSequence[n] == blastHit[n]:
							matchingSequence += guideSequence[n]
						else:
							break
						n += 1  

					# Guide match with PAM...
					# The only BLAST hits to worry about are those that include a PAM (or wobble PAM)...
					if blastHit[0:2] in ["CC", "CT", "TC", " C", " T"]: # Catching PAM and PAM wobble...
					#if blastHit[0:2] in ["CC", "CT", "TC"]: # Catching PAM and PAM wobble...

						# Have to extend into the genome to check PAM wobble...
						if blastHit[0] == " ":

							blastCheck = blastHitRaw
							gKeys = sorted(genome)
							for gKey in gKeys:
								result = genome[gKey].find(blastCheck)
								if result != -1:
									resultIndex = genome[gKey].index(blastCheck)
									blastCheckExtended = genome[gKey][resultIndex-1:resultIndex+len(blastCheck)]
									# print()
									# print(1, blastCheck)
									# print(2, blastCheckExtended)
									if blastCheckExtended[0] in "CT":

										# Conservative...
										if len(matchingSequence) > pamInclusionSequenceThreshold:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											pamInclusionsDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
											pamInclusions += 1
										# Super conservative...
										if len(matchingSequence) > 9:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											superConservativePamInclusionDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)

								result = genome[gKey].find(reverseComplement(blastCheck))
								if result != -1:
									resultIndex = genome[gKey].index(reverseComplement(blastCheck))
									blastCheckExtended = genome[gKey][resultIndex:resultIndex+len(blastCheck)+1]
									# print()
									# print("reversed...")
									# print(1, reverseComplement(blastCheck))
									# print(2, blastCheckExtended)
									if blastCheckExtended[-1] in "AG":

										# Conservative...
										if len(matchingSequence) > pamInclusionSequenceThreshold:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											pamInclusionsDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
											pamInclusions += 1
										# Super conservative...
										if len(matchingSequence) > 9:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											superConservativePamInclusionDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)

						else:

							# Conservative...
							if len(matchingSequence) > pamInclusionSequenceThreshold:
								# For monitoring PAM inclusions as we learn more about CRISPR...
								pamInclusionsDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
								pamInclusions += 1
							# Super conservative...
							if len(matchingSequence) > 9:
								# For monitoring PAM inclusions as we learn more about CRISPR...
								superConservativePamInclusionDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)

				else:

					# # Do not include the last two guide bases in the calculation...
					# n, matchingSequence = len(guideSequence) - 3, blastHit[-2:]
					# while n >= 0:
					# 	if guideSequence[0:-2][n] == blastHit[0:-2][n]:
					# 		matchingSequence += guideSequence[0:-2][n]
					# 	else:
					# 		break
					# 	n -= 1

					n, matchingSequence = len(guideSequence) - 1, ""
					while n >= 0:
						if guideSequence[n] == blastHit[n]:
							matchingSequence += guideSequence[n]
						else:
							break
						n -= 1

					# Guide match with PAM...
					# The only BLAST hits to worry about are those that include a PAM (or wobble PAM)...
					if blastHit[-2:] in ["GG", "GA", "AG", "G ", "A "]: # Catching PAM and PAM wobble...
					#if blastHit[-2:] in ["GG", "GA", "AG"]: # Catching PAM and PAM wobble...

						# Have to extend into the genome to check PAM wobble...
						if blastHit[-1] == " ":

							blastCheck = blastHitRaw
							gKeys = sorted(genome)
							for gKey in gKeys:
								result = genome[gKey].find(blastCheck)
								if result != -1:
									resultIndex = genome[gKey].index(blastCheck)
									#blastCheckExtended = genome[gKey][resultIndex-1:resultIndex+len(blastCheck)]
									blastCheckExtended = genome[gKey][resultIndex:resultIndex+len(blastCheck)+1]
									# print()
									# print(1, blastCheck)
									# print(2, blastCheckExtended)
									if blastCheckExtended[0] in "AG":

										# Conservative...
										if len(matchingSequence) > pamInclusionSequenceThreshold:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											pamInclusionsDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
											pamInclusions += 1
										# Super conservative...
										if len(matchingSequence) > 9:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											superConservativePamInclusionDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)

								result = genome[gKey].find(reverseComplement(blastCheck))
								if result != -1:
									resultIndex = genome[gKey].index(reverseComplement(blastCheck))
									#blastCheckExtended = genome[gKey][resultIndex:resultIndex+len(blastCheck)+1]
									blastCheckExtended = genome[gKey][resultIndex-1:resultIndex+len(blastCheck)]
									# print()
									# print("reversed...2")
									# print(1, reverseComplement(blastCheck))
									# print(2, blastCheckExtended)
									if blastCheckExtended[-1] in "CT":

										# Conservative...
										if len(matchingSequence) > pamInclusionSequenceThreshold:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											pamInclusionsDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
											pamInclusions += 1
										# Super conservative...
										if len(matchingSequence) > 9:
											# For monitoring PAM inclusions as we learn more about CRISPR...
											superConservativePamInclusionDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
						
						else:

							# Conservative...
							if len(matchingSequence) > pamInclusionSequenceThreshold:
								# For monitoring PAM inclusions as we learn more about CRISPR...
								pamInclusionsDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)
								pamInclusions += 1
							# Super conservative...
							if len(matchingSequence) > 9:
								# For monitoring PAM inclusions as we learn more about CRISPR...
								superConservativePamInclusionDictTmp[(guideSequence, line)] = (guideSequence, blastHit, line)

		i += 1

	return (safeGuides, unsafeGuides, pamInclusionsDict, superConservativePamInclusionDict)

def getOptiGuides(codons, guides, orfPlusSequence, maxPamCutGap, optiGuideInclusionsDict):


	# Find the optimal guide with minimum PAM inclusions and PAM gap...

	codonKeys, optiGuides, inclusionKeys = sorted(codons), {}, sorted(optiGuideInclusionsDict)
	for codonKey in codonKeys:

		keys, lastCodonBase, candidateGuidePool = sorted(guides), codonKey[-1], {}
		for key in keys:

			guide = guides[key] 
			guideIndex = orfPlusSequence.index(guide)
			if guide[0:2] == "CC":
				guideIndexCut = guideIndex + 6
			else:
				guideIndexCut = guideIndex + 17

			# Calculate the distance in bases between the Cas9 cut site and codon insertion point...
			gap = abs(lastCodonBase+1-guideIndexCut)
			# If the gap is small enough, calculation the number of PAM inclusions for the guide...
			if gap <= int(maxPamCutGap/2):
				nInclusions = 0
				for inclusionKey in inclusionKeys:
					if guide == inclusionKey[0]:
						nInclusions += 1
				candidateGuidePool[(nInclusions, gap, guide)] = key

		# The optimal guide...
		keys = sorted(candidateGuidePool)
		if keys:
			optiGuides[codonKey] = (candidateGuidePool[keys[0]], keys[0][2], keys[0][1], keys[0][0])
		# Or, NO optimal guide...
		else:
			optiGuides[codonKey] = None

	return optiGuides


##########################################################################################################################################################
# Functions for PAM-silencing...
##########################################################################################################################################################

def findPamSites(orfPlusSequence, orfStartIndex, orfEndIndex):

	pamsForward, pamsReverse = {}, {}
	while orfStartIndex < orfEndIndex - 2:
		pamaybe = orfPlusSequence[orfStartIndex:orfStartIndex+3]
		if "GG" == pamaybe[1:]:
			guide = orfPlusSequence[orfStartIndex-20:orfStartIndex+3]
			cutsite = orfStartIndex+3-5
			pamsForward[(cutsite, guide)] = guide
		if "CC" == pamaybe[0:2]:
			guide = orfPlusSequence[orfStartIndex:orfStartIndex+23]
			cutsite = orfStartIndex+6
			pamsReverse[(cutsite, guide)] = guide
		orfStartIndex += 1
	return (pamsForward, pamsReverse)

def tryToPamSilence(orfPlusSequence, orfStartIndex, orfEndIndex, guides):

	###############################################################################################
	# This function attempts to use silent mutations in the PAM site to cease Cas9 cutting...
	###############################################################################################

	#############################################
	# Primary PAM 5'--->3' top strand: 
	#	NGG or CCN
	# Secondary wobble PAMs 5'--->3' top strand:
	#	NGA or TCN
	#	NAG or CTN
	#############################################


	# Identify PAM sites that are (non)coincident with the orf reading frame...
	guideKeys = sorted(guides)
	inFrame, outFrame = {}, {}
	inFrameCodons = {}
	for guideKey in guideKeys:
		guide = guides[guideKey]
		guideStartIndex = orfPlusSequence.index(guide)
		guideEndIndex = guideStartIndex + len(guide)
		if guide[0:2] == "CC":
			frame = (guideStartIndex - orfStartIndex) % 3
			if not frame:
				inFrame[guideKey] = guide
				inFrameCodons[guideKey] = (guide[:3],  (guideStartIndex, guideStartIndex+3))
			else:
				outFrame[guideKey] = guide
		else:
			frame = ((guideStartIndex + len(guide)) - orfStartIndex) % 3
			if not frame:
				inFrame[guideKey] = guide
				inFrameCodons[guideKey] = (guide[-3:], ((guideStartIndex + len(guide))-3, (guideStartIndex + len(guide))))
			else:
				outFrame[guideKey] = guide

	# For PAM sites that are non-concident with the orf reading frame...
	# Identify the left and right codons that contribute to the PAM site...
	keys = sorted(outFrame)
	outFrameCodons = {}
	for key in keys:

		guide = outFrame[key]
		guideStartIndex = orfPlusSequence.index(guide)
		guideEndIndex = guideStartIndex + len(guide)

		# Guide begins with CCN...
		if guide[0:2] == "CC":

			frame = (guideStartIndex - orfStartIndex) / 3
			frame = str(frame)
			# 1st of 3 codon bases...out of frame by 1 base...
			if int(frame.split(".")[1][0]) == 3:
				lCodon = orfPlusSequence[guideStartIndex-1:guideStartIndex-1+3]
				rCodon = orfPlusSequence[guideStartIndex-1+3:guideStartIndex-1+6]
				outFrameCodons[key] = [(lCodon, (guideStartIndex-1,guideStartIndex-1+3)), (rCodon, (guideStartIndex-1+3,guideStartIndex-1+6), 3)]
			# 2nd of 3 codon bases...out of frame by 2 bases...
			if int(frame.split(".")[1][0]) == 6:
				lCodon = orfPlusSequence[guideStartIndex-2:guideStartIndex-2+3]
				rCodon = orfPlusSequence[guideStartIndex-2+3:guideStartIndex-2+6]
				outFrameCodons[key] = [(lCodon, (guideStartIndex-2, guideStartIndex-2+3)), (rCodon, (guideStartIndex-2+3, guideStartIndex-2+6), 6)]
		
		# Guide ends with NGG... 
		else:

			frame = ((guideStartIndex + len(guide)) - orfStartIndex) / 3
			frame = str(frame)
			# 1st of 3 codon bases...out of frame by 1 base...
			if int(frame.split(".")[1][0]) == 3:
				lCodon = orfPlusSequence[guideStartIndex+len(guide)-4:guideStartIndex+len(guide)-1]
				rCodon = orfPlusSequence[guideStartIndex+len(guide)-1:guideStartIndex+len(guide)+2]
				outFrameCodons[key] = [(lCodon, (guideStartIndex+len(guide)-4, guideStartIndex+len(guide)-1)), (rCodon, (guideStartIndex+len(guide)-1, guideStartIndex+len(guide)+2), 3)]
			# 2nd of 3 codon bases...out of frame by 2 bases...
			if int(frame.split(".")[1][0]) == 6:
				lCodon = orfPlusSequence[guideStartIndex+len(guide)-5:guideStartIndex+len(guide)-2]
				rCodon = orfPlusSequence[guideStartIndex+len(guide)-2:guideStartIndex+len(guide)+1]
				outFrameCodons[key] = [(lCodon, (guideStartIndex+len(guide)-5, guideStartIndex+len(guide)-2)), (rCodon, (guideStartIndex+len(guide)-2, guideStartIndex+len(guide)+1), 6)]

	# I may be able to trim some fat from here down...

	#############################################################################################
	# Try to silence inFrame PAMs...
	#############################################################################################

	keys = sorted(inFrameCodons)
	silencedGuides = {}
	for key in keys:

		guide = guides[key]
		codon = inFrameCodons[key][0]
		codonStartIndex = inFrameCodons[key][1][0]
		codonEndIndex = inFrameCodons[key][1][1]
		# In-frame CCN can only be PRO and can't be PAM silenced because all PRO codons begin with CC...
		if guide[0:2] == "CC":
			print("PAM SILENCING: Unsilenceable in-frame PAM.....from CCN", codon, byCodon[codon], guide)
		# In-frame NGG codons can be ARG, GLY, and TRP...
		# However, similar to PRO, the only TRP codon TTG cannot be PAM-silenced...
		else:
			res = byCodon[codon]
			possibleCodons = byRes[res]
			ranked, newCodon = sorted(possibleCodons)[::-1], None
			for rank in ranked:
				possibleCodon = possibleCodons[rank]
				# Avoid futile codons: codons that only change the N of the NGG...
				# Any codon that silences the GG of the NGG at the end of the guide...
				if possibleCodon[-2] != "G" or possibleCodon[-1] != "G":
					# Avoid wobble codons that may not provide the strongest silencing...
					#	NGA or TCN
					#	NAG or CTN
					if possibleCodon[-2:] not in ["GA", "AG"]:
						newCodon = possibleCodon
						break
					else:
						print("PAM SILENCING: Can't use new in-frame codon due to PAM wobble....", possibleCodon, byCodon[possibleCodon])

			if newCodon:
				silencedGuide = guide[:-3] + newCodon
				# newCodon is the same as the original codon...skip...
				if guide == silencedGuide:
					silenced = 0
					continue
				guideIndex = orfPlusSequence.find(guide)
				silencedOrf = orfPlusSequence[:guideIndex] + silencedGuide + orfPlusSequence[guideIndex+len(guide):]
				silencedGuide = markSilencers(silencedGuide, guide)
				silencedOrf = markSilencers(silencedOrf, orfPlusSequence)
				silencedGuides[key] = (silencedGuide, silencedOrf)

			# Any guide for which a newCodon cannot be generated, cannot be PAM-silenced...
			else:
				print("PAM SILENCING: Unsilenceable in-frame PAM.....from NGG", codon, byCodon[codon], guide)

	#############################################################################################
	# Try to silence outFrame PAMs...
	#############################################################################################

	# This is the more complicated scenario...
	# For guides that start with CCN, vary the left codon while the right codon is fixed...
	# For guides that end with NGG, fix the left codon and vary the right codon...
	keys = sorted(outFrameCodons)
	for key in keys:

		guide, silenced = guides[key], 0

		# Guide begins with CCN...
		if guide[0:2] == "CC":

			# Attempting to silence the left codon, fixed right codon...
			codons = outFrameCodons[key]
			rightCodonTuple = codons[1]
			leftCodonTuple = codons[0] 
			rightCodon = rightCodonTuple[0]
			leftCodon = leftCodonTuple[0]
			codonStartIndex = rightCodonTuple[1][0]
			codonEndIndex = rightCodonTuple[1][1]
			res = byCodon[leftCodon]
			possibleCodons = byRes[res]
			ranked, newCodon, replaceIndex = sorted(possibleCodons)[::-1], None, guide.find(rightCodon)
			for rank in ranked:
				possibleCodon = possibleCodons[rank]
				if possibleCodon[-1] != "C":
					# newCodon = possibleCodon
					# break
					# Avoid wobble codons that may not provide the strongest silencing...
					#	NGA or TCN
					#	NAG or CTN
					if possibleCodon[0:2] not in ["TC", "CT"]:
						newCodon = possibleCodon
						break
					else:
						print("PAM SILENCING: Can't use new in-frame codon due to PAM wobble....", possibleCodon, byCodon[possibleCodon])

			if newCodon:
				replaceIndex = guide.find(rightCodon)
				silencedGuide = newCodon[-replaceIndex:] + guide[replaceIndex:]
				# newCodon is the same as the original leftCodon...
				if guide == silencedGuide:
					silenced = 0
					continue
				if silencedGuide[:2].upper() == "CC":
					silenced = 0
					continue
				guideIndex = orfPlusSequence.find(guide)
				silencedOrf = orfPlusSequence[:guideIndex + replaceIndex - 3] + newCodon + silencedGuide[replaceIndex:] + orfPlusSequence[guideIndex+len(guide):]
				silencedGuide = markSilencers(silencedGuide, guide)
				silencedOrf = markSilencers(silencedOrf, orfPlusSequence)
				silencedGuides[key] = (silencedGuide, silencedOrf)
				silenced = 1

			# Any guide for which a newCodon cannot be generated, cannot be PAM-silenced...
			else:
				print("PAM SILENCING: Unsilenceable split-frame PAM.....from CCN", leftCodon, byCodon[leftCodon], rightCodon, byCodon[rightCodon], guide)

		# Guide ends with NGG...
		else:

			# Attempting to silence the left codon, fixed right codon...
			codons = outFrameCodons[key]
			rightCodonTuple = codons[1]
			leftCodonTuple = codons[0] 
			rightCodon = rightCodonTuple[0]
			leftCodon = leftCodonTuple[0]
			codonStartIndex = rightCodonTuple[1][0]
			codonEndIndex = rightCodonTuple[1][1]
			res = byCodon[leftCodon]
			possibleCodons = byRes[res]
			ranked, newCodon, replaceIndex = sorted(possibleCodons)[::-1], None, guide.rfind(leftCodon)
			for rank in ranked:
				possibleCodon = possibleCodons[rank]
				if possibleCodon[-1:] != "G":
					newCodon = possibleCodon
					break
			if newCodon:
				replaceIndex = guide.rfind(leftCodon)
				silencedGuide = guide[:replaceIndex] + newCodon + guide[replaceIndex+3:]
				# newCodon is the same as the original leftCodon...
				if guide == silencedGuide:
					silenced = 0
					continue
				if silencedGuide[-2:].upper() == "GG":
					silenced = 0
					continue
				guideIndex = orfPlusSequence.rfind(guide)
				silencedOrf = orfPlusSequence[:guideIndex] + silencedGuide[:replaceIndex] + newCodon + orfPlusSequence[guideIndex+len(silencedGuide[:replaceIndex] + newCodon):]
				silencedGuide = markSilencers(silencedGuide, guide)
				silencedOrf = markSilencers(silencedOrf, orfPlusSequence)
				silencedGuides[key] = (silencedGuide, silencedOrf)
				silenced = 1

			# Any guide for which a newCodon cannot be generated, cannot be PAM-silenced...
			else:
				print("PAM SILENCING: Unsilenceable split-frame PAM.....from NGG", leftCodon, byCodon[leftCodon], rightCodon, byCodon[rightCodon], guide)

	# Of those guides that cannot be PAM-silenced...
	# the majority are in-frame PAM sites with no avialable silencing codons....
	keys, unsilencedGuides = sorted(guides), {}
	for key in keys:
		if key not in silencedGuides:
			unsilencedGuides[key] = guides[key]

	return (silencedGuides, unsilencedGuides)

def testForWastedPamMutation(guide, silencedGuide):

	# Identify wasted silencing mutations in the N position of NGG or CCN PAM sites...

	wastedMutation, silencedGuide = 0, markSilencers(silencedGuide, guide)
	if guide[0:2] == "CC":
		if silencedGuide[2] in "ATGC":
			if silencedGuide[1] not in "ATGC":
				if silencedGuide[0] not in "ATGC":
					wastedMutation = 1
	else:
		if silencedGuide[-3] in "ATGC":
			if silencedGuide[-2] not in "ATGC":
				if silencedGuide[-1] not in "ATGC":
					wastedMutation = 1

	return wastedMutation

def guideSilence(orfPlusSequence, orfStartIndex, unsilencedGuides, mismatchThreshold=4):

	###############################################################################################
	# This function attempts to use silent mutations in the guide sequence to cease Cas9 cutting...
	# The code attempts to make the silent mutations as close as possible to the PAM site...
	###############################################################################################

	keys, silencedGuides = sorted(unsilencedGuides), {}
	for key in keys:

		silenced = 1
		guide = unsilencedGuides[key]

		# Guide begins with CCN...
		if guide[0:2] == "CC":

			# Based on guide frame, identify the sub-guide to silence...
			guideStartIndex = orfPlusSequence.index(guide)
			guideEndIndex = guideStartIndex + len(guide)
			frame = (guideStartIndex - orfStartIndex) / 3
			frame = str(frame)
			# 1st of 3 codon bases...out of frame by 1 base...
			if int(frame.split(".")[1][0]) == 3:
				trySilenceSubGuide = guide[2:]
			# 2nd of 3 codon bases...out of frame by 2 bases...
			elif int(frame.split(".")[1][0]) == 6:
				trySilenceSubGuide = guide[1:]
			else:
				trySilenceSubGuide = guide[3:]

			# Silence...
			i, totalMismatches, tmpTrySilenceSubGuide = 0, 0, trySilenceSubGuide
			while i < len(trySilenceSubGuide) and totalMismatches < mismatchThreshold:

				# Attempt to silence the guide using silent mutations...
				codon = trySilenceSubGuide[i:i+3]
				residue = byCodon[codon]
				codonOptions = sorted(byRes[residue]) # Rank by most commonly used codon...
				codonOptions.reverse()
				maxMismatches, maxTrySilenceSubGuide = 0, ""
				for codonOption in codonOptions:
					tmpTrySilenceSubGuideTest = tmpTrySilenceSubGuide[:i] + codonOption[1] + tmpTrySilenceSubGuide[i+3:]
					mismatches = countMismatches(tmpTrySilenceSubGuide, tmpTrySilenceSubGuideTest)
					
					# Do not allow for wasted silencing mutations in the N position of CCN...
					#########################################################################

					# Build the complete silenced guide...
					# 1st of 3 codon bases...out of frame by 1 base...
					if int(frame.split(".")[1][0]) == 3:
						silencedGuide = guide[0:2] + tmpTrySilenceSubGuideTest  
					# 2nd of 3 codon bases...out of frame by 2 bases...
					elif int(frame.split(".")[1][0]) == 6:
						silencedGuide = guide[0:1] + tmpTrySilenceSubGuideTest
					else:
						silencedGuide = guide[0:3] + tmpTrySilenceSubGuideTest		

					wmResult = testForWastedPamMutation(guide, silencedGuide)

					#########################################################################

					if not wmResult:
						if mismatches > maxMismatches:
							maxMismatches, maxTrySilenceSubGuide = mismatches, tmpTrySilenceSubGuideTest

				# Count the current number of mismatches between the original guide and current silenced guide...
				if maxMismatches:
					tmpTrySilenceSubGuide = maxTrySilenceSubGuide # Update the emerguing guide silenced sequence...
					deltaMismatch = countMismatches(trySilenceSubGuide, tmpTrySilenceSubGuide) # Update the emerging number of base mismatches...
					totalMismatches = deltaMismatch

				i += 3

			# Build the complete silenced guide...
			# 1st of 3 codon bases...out of frame by 1 base...
			if int(frame.split(".")[1][0]) == 3:
				silencedGuide = guide[0:2] + tmpTrySilenceSubGuide  
			# 2nd of 3 codon bases...out of frame by 2 bases...
			elif int(frame.split(".")[1][0]) == 6:
				silencedGuide = guide[0:1] + tmpTrySilenceSubGuide
			else:
				silencedGuide = guide[0:3] + tmpTrySilenceSubGuide

			# wmResult = testForWastedPamMutation(guide, silencedGuide)
			# if wmResult:
			# 	print()
			# 	print("wmResult -------------------------------")
			# 	print(wmResult)
			# 	print(guide)
			# 	print(markSilencers(silencedGuide, guide))

			# Build the complete silenced orf...
			silencedOrf = orfPlusSequence.replace(guide, silencedGuide)
			silencedGuide = markSilencers(silencedGuide, guide)
			if guide == silencedGuide:
				silenced = 0
				continue
			else:
				silencedOrf = markSilencers(silencedOrf, orfPlusSequence)
				silencedGuides[key] = (silencedGuide, silencedOrf)

		# Guide begins with NGG...
		else:

			# Based on guide frame, identify the sub-guide to silence...
			guideStartIndex = orfPlusSequence.index(guide)
			guideEndIndex = guideStartIndex + len(guide)
			frame = (guideEndIndex - orfStartIndex) / 3
			frame = str(frame)
			# 1st of 3 codon bases...out of frame by 1 base...
			if int(frame.split(".")[1][0]) == 3:
				trySilenceSubGuide = guide[:-1]
			# 2nd of 3 codon bases...out of frame by 2 bases...
			elif int(frame.split(".")[1][0]) == 6:
				trySilenceSubGuide = guide[:-2]
			else:
				trySilenceSubGuide = guide[:-3]

			# Silence...
			i, totalMismatches, tmpTrySilenceSubGuide = int(3*len(trySilenceSubGuide)/3), 0, trySilenceSubGuide
			while i > 0 and totalMismatches < mismatchThreshold:

				# Attempt to silence the guide using silent mutations...
				codon = trySilenceSubGuide[i-3:i]
				residue = byCodon[codon]
				codonOptions = sorted(byRes[residue]) # Rank by most commonly used codon...
				codonOptions.reverse()
				maxMismatches, maxTrySilenceSubGuide = 0, ""
				for codonOption in codonOptions:
					tmpTrySilenceSubGuideTest = tmpTrySilenceSubGuide[:i-3] + codonOption[1] + tmpTrySilenceSubGuide[i:]
					mismatches = countMismatches(tmpTrySilenceSubGuide, tmpTrySilenceSubGuideTest)

					# Do not allow for wasted silencing mutations in the N position of NGG...
					#########################################################################

					# Build the complete silenced guide...
					# 1st of 3 codon bases...out of frame by 1 base...
					if int(frame.split(".")[1][0]) == 3:
						silencedGuide = tmpTrySilenceSubGuideTest + guide[-1:]
					# 2nd of 3 codon bases...out of frame by 2 bases...
					elif int(frame.split(".")[1][0]) == 6:
						silencedGuide = tmpTrySilenceSubGuideTest + guide[-2:]
					else:
						silencedGuide = tmpTrySilenceSubGuideTest + guide[-3:]

					wmResult = testForWastedPamMutation(guide, silencedGuide)
					# if wmResult:
					# 	print()
					# 	print("wmResult -------------------------------")
					# 	print(wmResult)
					# 	print(guide)
					# 	print(markSilencers(silencedGuide, guide))

					#########################################################################

					if not wmResult:
						if mismatches > maxMismatches:
							maxMismatches, maxTrySilenceSubGuide = mismatches, tmpTrySilenceSubGuideTest

				# Count the current number of mismatches between the original guide and current silenced guide...
				if maxMismatches:
					tmpTrySilenceSubGuide = maxTrySilenceSubGuide # Update the emerguing guide silenced sequence...
					deltaMismatch = countMismatches(trySilenceSubGuide, tmpTrySilenceSubGuide) # Update the emerging number of base mismatches...
					totalMismatches = deltaMismatch

				i -= 3

			# Build the complete silenced guide...
			# 1st of 3 codon bases...out of frame by 1 base...
			if int(frame.split(".")[1][0]) == 3:
				silencedGuide = tmpTrySilenceSubGuide + guide[-1:]
			# 2nd of 3 codon bases...out of frame by 2 bases...
			elif int(frame.split(".")[1][0]) == 6:
				silencedGuide = tmpTrySilenceSubGuide + guide[-2:]
			else:
				silencedGuide = tmpTrySilenceSubGuide + guide[-3:]

			# wmResult = testForWastedPamMutation(guide, silencedGuide)
			# if wmResult:
			# 	print()
			# 	print("wmResult -------------------------------")
			# 	print(wmResult)
			# 	print(guide)
			# 	print(markSilencers(silencedGuide, guide))

			# Build the complete silenced orf...
			silencedOrf = orfPlusSequence.replace(guide, silencedGuide)
			silencedGuide = markSilencers(silencedGuide, guide)
			if guide == silencedGuide:
				silenced = 0
				continue
			else:
				silencedOrf = markSilencers(silencedOrf, orfPlusSequence)
				silencedGuides[key] = (silencedGuide, silencedOrf)

		# Not sure if this is possible...
		# Monitoring this situation since 2021.11.08...
		if not silenced:
			print("Could not be guide silenced.....", guide)

	return silencedGuides

def calculateScannableSequence(outputPath, geneName, orfSequence, orfPlusSequence, guides, maxPamCutGap):

    codonStartIndex = orfPlusSequence.find(orfSequence)
    if codonStartIndex == -1:
        raise ValueError("orfSequence not found within orfPlusSequence")

    keys = sorted(guides)  # key
    i, unscannableCodons = 0, []

    while i < len(keys) - 1:
        cut1 = keys[i][0]
        cut2 = keys[i + 1][0]

        if cut2 - cut1 > maxPamCutGap:
            gap, codons = "", []
            startIndex = cut1 + int(maxPamCutGap / 2)
            stopIndex = cut2 - int(maxPamCutGap / 2)

            while startIndex < stopIndex:
                gap += orfPlusSequence[startIndex]
                frame = (startIndex - codonStartIndex) % 3
                if not frame:
                    codons.append(int((startIndex - codonStartIndex) / 3))
                startIndex += 1

            # Store orfPlusSequence coords + gap + codon indices
            unscannableCodons.append(
                (cut1 + int(maxPamCutGap / 2), gap, cut2 - int(maxPamCutGap / 2), codons)
            )

        i += 1

    # Start with the raw ORF sequence
    scannableSequence = orfSequence
    orf_start = codonStartIndex  # ORF start index in orfPlusSequence

    # Blank out unscannable regions, but using ORF-relative coordinates
    for start_plus, gap, stop_plus, codon_list in unscannableCodons:
        # Convert from orfPlusSequence coordinates to ORF coordinates
        start = start_plus - orf_start
        stop = stop_plus - orf_start

        # Skip if the gap is completely outside the ORF
        if stop <= 0 or start >= len(scannableSequence):
            continue

        # Clamp to ORF bounds
        start = max(0, start)
        stop = min(len(scannableSequence), stop)

        # Replace the ORF region with spaces
        scannableSequence = (
            scannableSequence[:start]
            + " " * (stop - start)
            + scannableSequence[stop:]
        )

    # Count blanks and report fraction scannable
    blanks = sum(1 for c in scannableSequence if c == " ")

    # Compute raw fraction without rounding up to 1.0
    raw_fraction = 1 - blanks / len(scannableSequence)

    # Represent with 3 decimal places, but avoid printing as 1.000 unless exact
    fraction_scannable = float(f"{raw_fraction:.3f}")
    if blanks > 0 and fraction_scannable >= 1.0:
        # Ensure it does not appear as 1.0 if there are any blanks
        # Use the next representable value below 1.0 at this precision
        fraction_scannable = 0.999

    with open(outputPath + geneName + "-scannableSequence.txt", "w") as f:
        f.write(
            "Fraction of total PAM-scannable sequence for the ORF: "
            + str(fraction_scannable)
            + "\n"
        )
        print(
            "Fraction of total PAM-scannable sequence for the ORF: "
            + str(fraction_scannable)
        )

        scannableSequence_fasta = fasta(scannableSequence)
        f.write(scannableSequence_fasta + "\n")

        if unscannableCodons:
            f.write("Affected codons:\n")
            for start_plus, gap, stop_plus, codon_list in unscannableCodons:
                f.write(gap + " " + str(codon_list) + "\n")

def createPrimerOrder(path, microplateWells, geneName, guidePrimers, insertPrimers):

	# Prepare guide primers...
	guidePrimersKeys = sorted(guidePrimers)
	i = 0
	primerOrder = {}
	for guidePrimersKey in guidePrimersKeys:	
		primerOrder[i] = (guidePrimersKey, guidePrimers[guidePrimersKey])
		i += 1

	# Prepare insertion primers...
	insertPrimersKeys = sorted(insertPrimers)
	for insertPrimersKey in insertPrimersKeys:
		primerOrder[i] = (insertPrimersKey, insertPrimers[insertPrimersKey])
		i += 1

	# Create microplate instance...
	from pam_scanning.plates import Microplate
	primerOrderMicroplate = Microplate()
	if microplateWells == 96:
		primerOrderMicroplate.create(8, 12)
	else:
		primerOrderMicroplate.create(16, 24)

	import openpyxl
	primerOrderExcel = openpyxl.Workbook()
	sheet = primerOrderExcel["Sheet"]
	primerOrderExcel.remove(sheet)
	sheet = primerOrderExcel.create_sheet("Plate Order 1")
	cell = sheet.cell(row=1, column=1)
	cell.value = "Well Position"
	cell = sheet.cell(row=1, column=2)
	cell.value = "Name"
	cell = sheet.cell(row=1, column=3)
	cell.value = "Sequence"
	primerPlateMapping = {}
	keys = sorted(primerOrder)
	j, p, r = 0, 1, 2
	for key in keys:

		well = primerOrderMicroplate.plateArrayTransposed[j]
		primer = primerOrder[key]
		primerName = primer[0]
		primerSequence = primer[1]

		cell = sheet.cell(row=r, column=1)
		cell.value = well
		cell = sheet.cell(row=r, column=2)
		cell.value = primerName
		cell = sheet.cell(row=r, column=3)
		cell.value = primerSequence

		primerPlateMapping[primerName] = (p, j)

		j += 1
		r += 1

		if j == len(primerOrderMicroplate.plateArrayTransposed):
			p += 1
			sheet = primerOrderExcel.create_sheet("Plate Order " + str(p))
			cell = sheet.cell(row=1, column=1)
			cell.value = "Well Position"
			cell = sheet.cell(row=1, column=2)
			cell.value = "Name"
			cell = sheet.cell(row=1, column=3)
			cell.value = "Sequence"
			j, r = 0, 2

	primerOrderExcel.save(path + geneName + "-primerOrder.xlsx")
