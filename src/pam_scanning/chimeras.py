"""High-level PAM-scanning driver.

:func:`pamscan` is the single entry point shared by the command-line interface
(:mod:`pam_scanning.cli`) and the Tkinter GUI (:mod:`pam_scanning.gui`). It takes
the run parameters as keyword arguments, runs the full pipeline implemented in
:mod:`pam_scanning.library`, and writes a time-stamped results directory.

Each 5'/3' flank may be supplied either as a FASTA file (``flank5_file_path`` /
``flank3_file_path``) or as a literal sequence (``flank5_sequence`` /
``flank3_sequence``), but not both for the same side. Sequences that were typed
or pasted rather than read from a file are written into the QC directory as
FASTA so every run remains fully reproducible from its own output.
"""

import re


# Sentinel written by the GUI (and defaulted by the CLI) for an unset path.
NOT_SELECTED = "No file selected"

# The pipeline operates on unambiguous DNA; 'N' is tolerated in flanking context.
_DNA_BASES = frozenset("ACGTN")


def _is_provided(value):
	"""True when a path/sequence kwarg holds a real value rather than a sentinel."""
	return bool(value) and value != NOT_SELECTED


def default_codon_table_path():
	"""Return the path to the yeast codon table bundled as package data."""
	from importlib import resources

	return str(resources.files("pam_scanning") / "data" / "codon_tables" / "yeast_64_1_1_all_nuclear.cusp.txt")


# The bundled default genome (S. cerevisiae BY4741) ships gzipped to keep the
# package small; it is decompressed once into a user cache on first use.
DEFAULT_GENOME_NAME = "BY4741_Toronto_2012.fsa"


def default_genome_path():
	"""Return a path to the bundled yeast genome FASTA, decompressing it if needed.

	The genome is stored gzipped as package data and expanded once into
	``~/.pam_scanning/genome`` (reused on later runs). Returns the path to the
	decompressed FASTA, ready for the off-target BLAST genome parser.
	"""
	import gzip
	import os
	import shutil
	from importlib import resources

	cache_dir = os.path.join(os.path.expanduser("~"), ".pam_scanning", "genome")
	cache_path = os.path.join(cache_dir, DEFAULT_GENOME_NAME)
	source = resources.files("pam_scanning") / "data" / "genomes" / (DEFAULT_GENOME_NAME + ".gz")

	# Reuse the cache only if it matches the packaged file's size (guards against a
	# truncated earlier expansion or a package update).
	with resources.as_file(source) as gz_path:
		if not os.path.exists(cache_path) or os.path.getsize(cache_path) == 0:
			os.makedirs(cache_dir, exist_ok=True)
			tmp = cache_path + ".tmp"
			with gzip.open(gz_path, "rb") as fin, open(tmp, "wb") as fout:
				shutil.copyfileobj(fin, fout)
			os.replace(tmp, cache_path)
	return cache_path


def parse_sequence_text(text, label="sequence"):
	"""Parse typed or pasted sequence text into a bare uppercase DNA string.

	Accepts a raw sequence or a whole pasted FASTA record: '>' header lines are
	dropped, and whitespace plus any base-numbering digits (as pasted from
	sequence viewers) are removed. Raises :class:`ValueError` if the result is
	empty or holds anything other than A, C, G, T, or N.
	"""
	body = [line for line in str(text).splitlines() if not line.lstrip().startswith(">")]
	sequence = re.sub(r"[\s\d]", "", "".join(body)).upper()
	if not sequence:
		raise ValueError("The %s is empty." % label)
	invalid = sorted(set(sequence) - _DNA_BASES)
	if invalid:
		raise ValueError(
			"The %s contains non-DNA character(s): %s (expected A, C, G, T, or N)."
			% (label, ", ".join(invalid))
		)
	return sequence


def _read_fasta_sequence(path):
	"""Read a FASTA file into a single uppercase sequence string.

	Mirrors the original inline reader: header ('>') lines are skipped and the
	SnapGene '\\r\\n' line ending is handled.
	"""
	seq = ""
	sequenceInput = open(path, "r")
	for line in sequenceInput.readlines():
		if line[0] == ">":
			continue
		if "\r\n" in line:
			sLine = line.split("\r\n")  # Necessary becuase this is how SnapGene exports FASTA files...
		else:
			sLine = line.split("\n")
		if sLine:
			seq += sLine[0]
	sequenceInput.close()
	return seq.upper()  # Initial sequence must be uppercase for functions to work properly...


def _resolve_flank(sequence, path, label):
	"""Return the flank sequence, from a literal sequence or a FASTA file."""
	if _is_provided(sequence) and _is_provided(path):
		raise ValueError(
			"Supply the %s flank either as a FASTA file or as a sequence, not both." % label
		)
	if _is_provided(sequence):
		return parse_sequence_text(sequence, "%s flank sequence" % label)
	return _read_fasta_sequence(path)


def _write_flank_qc(qcPath, geneName, label, path, sequence):
	"""Preserve the flank in QC: copy its FASTA, or write the entered sequence as one."""
	from os import sep
	from shutil import copyfile
	from pam_scanning.library import fasta

	if _is_provided(path):
		copyfile(path, qcPath + path.split(sep)[-1])
		return
	output = open(qcPath + geneName + "-" + label + ".fa", "w")
	output.write("> " + geneName + " | " + label + " (entered sequence)\n")
	output.write(fasta(sequence))
	output.close()


def _build_summary(gene, n_codons, n_bp, fraction, designed_sites, requested_sites,
                   unscannable_sites, n_guide_primers, n_insert_primers, n_warnings, output_dir):
	"""Compose the end-of-run summary shown in the progress console."""
	rule = "=" * 64
	rows = [
		("ORF length", "%d codons (%d bp)" % (n_codons, n_bp)),
		("PAM-scannable", "%.1f%% of the ORF" % (100.0 * fraction)),
		("Insertion sites", "%d of %d requested codon(s) have a guide" % (designed_sites, requested_sites)),
		("Unscannable sites", "%d" % unscannable_sites),
		("Guide primers", "%d" % n_guide_primers),
		("Insertion primers", "%d" % n_insert_primers),
		("PAM-inclusion warnings", "%d" % n_warnings),
		("Output", output_dir),
	]
	body = "\n".join("  %-24s %s" % (label + " " + "." * (22 - len(label)), value)
	                 for label, value in rows)
	return "\n%s\n  PAM-scan summary: %s\n%s\n%s\n%s" % (rule, gene, rule, body, rule)


def pamscan(**kwargs):

	# Set **kwargs

	orf_file_path = kwargs['orf_file_path']
	# Each flank arrives either as a FASTA path or as a literal sequence.
	flank5_file_path = kwargs.get('flank5_file_path', NOT_SELECTED)   # 100 bp upstream of the ATG (the "-" side)
	flank3_file_path = kwargs.get('flank3_file_path', NOT_SELECTED)   # 100 bp downstream of the stop (the "+" side)
	flank5_sequence = kwargs.get('flank5_sequence')
	flank3_sequence = kwargs.get('flank3_sequence')
	local_genome_file_path = kwargs['local_genome_file_path']
	codon_table_file_path = kwargs.get('codon_table_file_path', 'No file selected')
	codon_selection_file_path = kwargs['codon_selection_file_path']

	geneName = kwargs['geneName']
	localBlastDb = kwargs['localBlastDb']
	guidePrimerForwardSuffix = kwargs['guidePrimerForwardSuffix']
	insertPrimerForwardSuffix = kwargs['insertPrimerForwardSuffix']
	insertPrimerReverseSuffix = kwargs['insertPrimerReverseSuffix']

	primerLength = kwargs['primerLength']
	maxPamCutGap = kwargs['maxPamCutGap']
	codonsSamplingGap = kwargs['codonsSamplingGap']
	pamInclusionThreshold = kwargs['pamInclusionThreshold']
	pamInclusionSequenceThreshold = kwargs['pamInclusionSequenceThreshold'] 
	codonsSamplingGap = kwargs['codonsSamplingGap']

	outputPath = kwargs['outputPath']

	if orf_file_path == 'No file selected':
		print("ORF file is required")
		return 0

	if not _is_provided(flank5_file_path) and not _is_provided(flank5_sequence):
		print("5' flank (100 bp upstream of ATG) is required: give a FASTA file or a sequence")
		return 0

	if not _is_provided(flank3_file_path) and not _is_provided(flank3_sequence):
		print("3' flank (100 bp downstream of stop) is required: give a FASTA file or a sequence")
		return 0

	if local_genome_file_path == 'No file selected':
		print("Local genome file is required")
		return 0

	# No codon table supplied: fall back to the bundled yeast codon table.
	if not codon_table_file_path or codon_table_file_path == 'No file selected':
		codon_table_file_path = default_codon_table_path()

	# No BLAST database supplied: build one (once, cached) from the genome itself, so
	# the genome is the single input and the database always matches what is scanned.
	# An explicit database name/path, when given, overrides this.
	if not _is_provided(localBlastDb):
		from pam_scanning import blast_setup
		localBlastDb = blast_setup.ensure_blast_db(local_genome_file_path)

	########################################################################################
	# Primer-order layout settings...
	########################################################################################

	# Microplate dimensions for the primer order...Options 96 or 384...
	microplateDimensions = 96

	########################################################################################
	# Get dependencies...
	########################################################################################

	from pam_scanning.library import fasta
	from pam_scanning.library import byRes, byCodon, singleLetterAA, reverseComplement
	from pam_scanning.library import findPamSites, tryToPamSilence, guideSilence, blastGuides, createPrimerOrder
	from pam_scanning.library import getOptiGuides, calculateScannableSequence
	from pam_scanning.library import set_codon_table
	import openpyxl

	set_codon_table(codon_table_file_path)

	# If they exist, collect user-provided codon selections...
	selectCodons = []
	if codon_selection_file_path and codon_selection_file_path != "No file selected":
		workbook = openpyxl.load_workbook(codon_selection_file_path)
		sheet = workbook["Mutations"]
		rows = sheet.max_row
		r = 2
		while r <= rows:
			cell = sheet.cell(row=r, column=1)
			residueNumber = cell.value
			# Necessary to skip previously occupied workbook cells that have been deleted... 
			if residueNumber == None:
				r += 1
				continue
			cell = sheet.cell(row=r, column=2)
			originalResidue = cell.value
			cell = sheet.cell(row=r, column=3)
			mutatedResidue = cell.value
			selectCodons.append((residueNumber, originalResidue, mutatedResidue))
			r += 1
	selectCodons.sort()

	# if not selectCodons:
	# 	print("No mutations were provided...abort PAM scan.")
	# 	raise SystemExit

	########################################################################################
	# Get DNA sequences...
	########################################################################################

	# A) Get the ORF sequence (ATG to stop) and the 100 bp genomic flanks on each side.
	# Flanks come from a FASTA file or straight from an entered sequence.
	orfSequence = _read_fasta_sequence(orf_file_path)
	flank5Sequence = _resolve_flank(flank5_sequence, flank5_file_path, "5'")   # 100 bp upstream of the ATG
	flank3Sequence = _resolve_flank(flank3_sequence, flank3_file_path, "3'")   # 100 bp downstream of the stop

	# B) Assemble the ORF-plus-context sequence: 5' flank + ORF + 3' flank. This is the
	# sequence the whole algorithm scans; building it from explicit flanks lets us reach
	# guide/primer positions at the very ends of the ORF.
	orfPlusSequence = flank5Sequence + orfSequence + flank3Sequence

	########################################################################################
	# Make a time-stamped directory for the Pamscanning output....
	########################################################################################

	from os import mkdir, listdir, sep
	from os.path import exists
	from shutil import copyfile
	from time import strftime
	timeStamp = strftime("%Y.%m.%d-%H.%M.%S")

	# Root calculation results path...
	# outputPath = ".." + sep + "calculationResults" + sep
	outputPath += sep
	print("Writing results to:", outputPath)
	if not exists(outputPath):
		mkdir(outputPath)
	outputPath += geneName + "-chimera-insertions-" + timeStamp + sep
	if not exists(outputPath):
		mkdir(outputPath)

	# QC path...
	qcPath = outputPath + "QC" + sep
	if not exists(qcPath):
		mkdir(qcPath)

	# BLAST+ path...
	blastPath = outputPath + "BLAST+" + sep
	if not exists(blastPath):
		mkdir(blastPath)

	# ORDER path...
	orderPath = outputPath + "ORDER" + sep
	if not exists(orderPath):
		mkdir(orderPath)

	if not exists(qcPath + "solutionsFasta" + sep):
		mkdir(qcPath + "solutionsFasta" + sep)
	copyfile(orf_file_path, qcPath + orf_file_path.split(sep)[-1])

	########################################################################################
	# Copy the flanks and write the assembled ORF-plus-context sequence for QC checks...
	########################################################################################

	_write_flank_qc(qcPath, geneName, "flank5", flank5_file_path, flank5Sequence)
	_write_flank_qc(qcPath, geneName, "flank3", flank3_file_path, flank3Sequence)
	orfPlusOut = open(qcPath + geneName + "-orfPlusContext.fa", "w")
	orfPlusOut.write("> " + geneName + " | ORF + 5' and 3' flanks (assembled)\n")
	orfPlusOut.write(fasta(orfPlusSequence))
	orfPlusOut.close()

	########################################################################################
	# Set codon information...
	########################################################################################

	orfStartIndex = orfPlusSequence.index(orfSequence)
	orfEndIndex = orfPlusSequence.index(orfSequence) + len(orfSequence)
	orfOffset = orfPlusSequence.index(orfSequence)
	i, allCodons = orfPlusSequence.index(orfSequence), {}
	while i < orfEndIndex:
		codon = orfPlusSequence[i:i+3]
		allCodons[(i, i+1, i+2)] = (codon, int(((i-orfOffset)/3) + 1))
		i += 3

	# Filter codons by A) selectCodons or B) codonSamplingGap if set...
	codonKeys = sorted(allCodons)
	# A) Filter by selectCodons...
	filteredCodons = {}
	if selectCodons:
		for selectCodon in selectCodons:
			residueNumber = selectCodon[0]
			for codonKey in codonKeys:
				if residueNumber == allCodons[codonKey][1]:
					filteredCodons[codonKey] = allCodons[codonKey]
					break
	# B) Filter by codonSamplingGap
	else:
		i = 0
		while i < len(codonKeys):
			codonKey = codonKeys[i]
			filteredCodons[codonKey] = allCodons[codonKey]
			i += codonsSamplingGap
	codons = filteredCodons
	codonKeys = sorted(codons)

	########################################################################################
	# Identify and attempt to silence PAM sites...
	########################################################################################

	print("Finding PAM sites and candidate guides...")
	result = findPamSites(orfPlusSequence, orfStartIndex, orfStartIndex + len(orfSequence))
	guidesF, guidesR = result[0], result[1]
	guides = guidesF
	guides.update(guidesR)
	print("Silencing PAM sites...")
	result = tryToPamSilence(orfPlusSequence, orfStartIndex, orfEndIndex, guides)
	silencedGuides, unsilencedGuides = result[0], result[1]

	########################################################################################
	# Attempt to GUIDE silence any guide that cannot be PAM silenced...
	########################################################################################

	print("Guide-silencing any remaining guides...")
	guideSilencers = guideSilence(orfPlusSequence, orfStartIndex, unsilencedGuides)
	silencedGuides.update(guideSilencers)
	for key in silencedGuides:
		if key in unsilencedGuides:
			del unsilencedGuides[key]
	# Update guides to only silenceable guides...
	tmpGuides = {}
	keys = sorted(silencedGuides)
	for key in keys:
		tmpGuides[key] = guides[key]
	guides = tmpGuides

	########################################################################################
	# BLAST guides against the reference genome to avoid off-target Cas9 cutting...
	########################################################################################

	result = blastGuides(blastPath, geneName, guides, localBlastDb, local_genome_file_path, pamInclusionSequenceThreshold=pamInclusionSequenceThreshold, pamInclusionThreshold=pamInclusionThreshold)
	safeGuides, unsafeGuides, pamInclusionsDict, superConservativePamInclusionDict = result[0], result[1], result[2], result[3]
	keys, tmpGuides, tmpSilencedGuides = sorted(safeGuides), {}, {}
	for key in keys:
		# Remove any guides that could lead to off-target cutting...
		tmpGuides[key] = guides[key]
		tmpSilencedGuides[key] = silencedGuides[key]
	guides, silencedGuides = tmpGuides, tmpSilencedGuides

	# Identify the optimal safeguide for each codon (post BLAST verification)....
	# This information is used to calculate total PAM-scannable sequence for the ORF...
	print("Selecting the optimal guide for each codon...")
	optiGuidesAllCodons = getOptiGuides(allCodons, guides, orfPlusSequence, maxPamCutGap, pamInclusionsDict)
	keys, tmpGuides, tmpSilencedGuides = sorted(optiGuidesAllCodons), {}, {}
	for key in keys:
		if not optiGuidesAllCodons[key]:
			continue
		guideKey = optiGuidesAllCodons[key][0]
		guideSeq = optiGuidesAllCodons[key][1]
		guideCut = optiGuidesAllCodons[key][2]
		tmpGuides[guideKey] = guides[guideKey]
		tmpSilencedGuides[guideKey] = silencedGuides[guideKey]
	guides, silencedGuides = tmpGuides, tmpSilencedGuides

	# Identify the optimal safeguide for each codon of interest (post BLAST verification)....
	optiGuides = getOptiGuides(codons, guides, orfPlusSequence, maxPamCutGap, pamInclusionsDict)
	keys, noOptiGuideCodonSet = sorted(optiGuides), {}
	for key in keys:
		if not optiGuides[key]:
			noOptiGuideCodonSet[key] = None

	########################################################################################
	# Make and bundle necessary guide and payload primers...
	########################################################################################

	wb = openpyxl.Workbook()
	sheet = wb["Sheet"]
	wb.remove(sheet)
	sheet = wb.create_sheet("Guide Solutions By Row")
	cell = sheet.cell(row=1, column=1)
	cell.value = "Codon"
	cell = sheet.cell(row=1, column=2)
	cell.value = "Cut @ Base"
	cell = sheet.cell(row=1, column=3)
	cell.value = "PAM Cut Gap"
	cell = sheet.cell(row=1, column=4)
	cell.value = "PAM Inclusions"
	cell = sheet.cell(row=1, column=5)
	cell.value = "Original Guide"
	cell = sheet.cell(row=1, column=6)
	cell.value = "Silenced Guide"
	cell = sheet.cell(row=1, column=7)
	cell.value = "Primer"
	cell = sheet.cell(row=1, column=8)
	cell.value = "Primer Name"
	cell = sheet.cell(row=1, column=9)
	cell.value = "Primer Guide"

	codonKeys, row, primerOrderGuides, primerOrderInserts = sorted(optiGuides), 2, {}, {}
	for codonKey in codonKeys:

		if codonKey in noOptiGuideCodonSet:
			continue

		# Get codon information...
		codon = codons[codonKey]
		codonBases = codon[0]
		codonNumber = codon[1]

		# Build primers...
		guideKey = optiGuides[codonKey][0]
		guideSeq = optiGuides[codonKey][1].lower()
		guideGap = optiGuides[codonKey][2]
		guideInclusions = optiGuides[codonKey][3]
		silencedGuide = silencedGuides[guideKey][0]
		silencedOrf = silencedGuides[guideKey][1]
		primerLengthF = primerLength - len(insertPrimerForwardSuffix)
		primerLengthR = primerLength - len(insertPrimerReverseSuffix)
		leftHomologyArm = silencedOrf[codonKey[0]-primerLengthF+3:codonKey[0]+3]
		rightHomologyArm = silencedOrf[codonKey[0]+3:codonKey[0]+3+primerLengthR]
		primerF = leftHomologyArm + insertPrimerForwardSuffix
		primerR = reverseComplement(rightHomologyArm) + insertPrimerReverseSuffix
		primerNameGuide = str(guideKey[0]) + ".gF"
		primerNameInsertionF = str(guideKey[0]) + ".iF." + str(codonNumber)
		primerNameInsertionR = str(guideKey[0]) + ".iR." + str(codonNumber)
		if guideSeq[-2:] == "gg":
			primerOrderGuides[primerNameGuide] = guideSeq[:-3] + guidePrimerForwardSuffix
		else:
			primerOrderGuides[primerNameGuide] = reverseComplement(guideSeq)[:-3] + guidePrimerForwardSuffix
		primerOrderInserts[primerNameInsertionF] = primerF
		primerOrderInserts[primerNameInsertionR] = primerR

		# Write .fa file for viewing in SnapGene...
		fastaHeader = "> "
		fastaHeader += "Codon: " + str(codonNumber)
		fastaHeader += " | Original guide: " + guideSeq 
		fastaHeader += " | Silenced guide: " + silencedGuide
		fastaHeader += " | Left homology arm: " + leftHomologyArm
		fastaHeader += " | Right homology arm: " + rightHomologyArm 
		fastaHeader += " | Gap: " + str(guideGap)  
		fastaHeader += " | Inclusions: " + str(guideInclusions) + "\n"
		output = open(qcPath + "solutionsFasta" + sep + geneName + "-" + str(codonNumber) + ".fa", "w")
		output.write(fastaHeader)
		output.write(fasta(silencedOrf))
		output.close()

		# Add solution to the .xls solution workbook...
		# Forward...
		cell = sheet.cell(row=row, column=1)
		cell.value = codonNumber
		cell = sheet.cell(row=row, column=2)
		cell.value = guideKey[0]
		cell = sheet.cell(row=row, column=3)
		cell.value = guideGap
		cell = sheet.cell(row=row, column=4)
		cell.value = guideInclusions
		cell = sheet.cell(row=row, column=5)
		cell.value = guideSeq
		cell = sheet.cell(row=row, column=6)
		cell.value = silencedGuide
		cell = sheet.cell(row=row, column=7)
		cell.value = primerF
		cell = sheet.cell(row=row, column=8)
		cell.value = primerNameInsertionF
		cell = sheet.cell(row=row, column=9)
		cell.value = primerNameGuide

		# Reverse...
		cell = sheet.cell(row=row+1, column=1)
		cell.value = codonNumber 
		cell = sheet.cell(row=row+1, column=2)
		cell.value = guideKey[0]
		cell = sheet.cell(row=row+1, column=3)
		cell.value = guideGap
		cell = sheet.cell(row=row+1, column=4)
		cell.value = guideInclusions
		cell = sheet.cell(row=row+1, column=5)
		cell.value = guideSeq
		cell = sheet.cell(row=row+1, column=6)
		cell.value = silencedGuide
		cell = sheet.cell(row=row+1, column=7)
		cell.value = primerR
		cell = sheet.cell(row=row+1, column=8)
		cell.value = primerNameInsertionR
		cell = sheet.cell(row=row+1, column=9)
		cell.value = primerNameGuide

		row += 2

	wb.save(qcPath + geneName + "-guideSolutions.xlsx")

	########################################################################################
	# BLAST guides against the reference genome...
	# At the this point, the remaining guides are safe and avoid off-target Cas9 cutting...
	# The purpose of this code is to generate the x-guidesBlastResult-final.txt
	########################################################################################

	optiGuidesForFinalBlast = {}
	keys = sorted(optiGuides)
	for key in keys:

		if not optiGuides[key]:
			continue

		guideKey = optiGuides[key][0]
		guideSequence = guideKey[1]
		optiGuidesForFinalBlast[guideKey] = guideSequence

	result = blastGuides(blastPath, geneName, optiGuidesForFinalBlast, localBlastDb, local_genome_file_path, pamInclusionSequenceThreshold=pamInclusionSequenceThreshold, pamInclusionThreshold=pamInclusionThreshold, final=1)
	pamInclusionsDict, superConservativePamInclusionDict = result[2], result[3]

	########################################################################################
	# Create the primer order...
	########################################################################################

	print("Writing primer order and QC files...")
	createPrimerOrder(orderPath, microplateDimensions, geneName, primerOrderGuides, primerOrderInserts)

	########################################################################################
	# Report PAM inclusions warning for potential off-target cutting...
	########################################################################################

	codonKeys, writeInclusions, writeInclusionsSuper = sorted(optiGuides), {}, {}
	for codonKey in codonKeys:

		if codonKey in noOptiGuideCodonSet:
			continue

		# Get codon information...
		codon = codons[codonKey]
		codonBases = codon[0]
		codonNumber = codon[1]
		optiGuide = optiGuides[codonKey][0][1]

		# Conservative PAM inclusions...
		keys = sorted(pamInclusionsDict)
		for key in keys:
			inclusions = pamInclusionsDict[key]
			inclusionGuide = inclusions[0]
			if optiGuide == inclusionGuide:
				print("Warning: Some guides will contain potential PAM inclusions...", inclusions)
				writeInclusions[(codonNumber, key)] = pamInclusionsDict[key]

		# Super conservative PAM inclusions...
		keys = sorted(superConservativePamInclusionDict)
		for key in keys:
			inclusions = superConservativePamInclusionDict[key]
			inclusionGuide = inclusions[0]
			if optiGuide == inclusionGuide:
				writeInclusionsSuper[(codonNumber, key)] = superConservativePamInclusionDict[key]

	if writeInclusions:

		# Warning path...
		warningPath = outputPath + "WARNINGS/"
		if not exists(warningPath):
			mkdir(warningPath)

		f = open(warningPath + geneName + "-pamInclusionWarnings.txt", "w")
		sFormat = "%-6s | %-25s | %-25s | %-50s\n"
		f.write(sFormat % ("Codon", "Guide", "BLAST hit", "BLAST result"))
		sFormat = "%-6i | %-25s | %-25s | %-50s\n"
		for key in sorted(writeInclusions):
			inclusions = writeInclusions[key]
			f.write(sFormat % (key[0], inclusions[0], inclusions[1], inclusions[2]))
		f.close()

	# Don't write the super conservative inclusion file if it is the same as the parameter-defined inclusion file...
	if writeInclusionsSuper and len(superConservativePamInclusionDict) != len(pamInclusionsDict):

		# Warning path...
		warningPath = outputPath + "WARNINGS/"
		if not exists(warningPath):
			mkdir(warningPath)

		f = open(warningPath + geneName + "-pamInclusionWarnings-superConservative.txt", "w")
		sFormat = "%-6s | %-25s | %-25s | %-50s\n"
		f.write(sFormat % ("Codon", "Guide", "BLAST hit", "BLAST result"))
		sFormat = "%-6i | %-25s | %-25s | %-50s\n"
		for key in sorted(writeInclusionsSuper):
			inclusions = writeInclusionsSuper[key]
			f.write(sFormat % (key[0], inclusions[0], inclusions[1], inclusions[2]))
		f.close()

	if noOptiGuideCodonSet:

		# Warning path...
		warningPath = outputPath + "WARNINGS/"
		if not exists(warningPath):
			mkdir(warningPath)

		f = open(warningPath + geneName + "-unscannableCodons.txt", "w")
		keys = sorted(noOptiGuideCodonSet)
		for key in keys:
			# Get codon information...
			codon = codons[key]
			codonBases = codon[0]
			codonNumber = codon[1]
			f.write(str(codonNumber) + "\n")
		f.close()

	########################################################################################
	# Report the total PAM-scannable sequence of the ORF for reference...
	########################################################################################

	fraction_scannable, scannableSequence = calculateScannableSequence(
		qcPath, geneName, orfSequence, orfPlusSequence, safeGuides, maxPamCutGap)

	########################################################################################
	# Summary report (shown in the console) + PAM-scannability plot (QC + console)...
	########################################################################################

	designedSites = sum(1 for k in optiGuides if optiGuides[k])
	summary = _build_summary(
		geneName, len(orfSequence) // 3, len(orfSequence), fraction_scannable,
		designedSites, len(codons), len(noOptiGuideCodonSet),
		len(primerOrderGuides), len(primerOrderInserts), len(writeInclusions), outputPath)
	print(summary)
	with open(qcPath + geneName + "-summary.txt", "w") as summaryFile:
		summaryFile.write(summary + "\n")

	plot_png = None
	try:
		from pam_scanning.plots import plot_scannable_positions
		plot_png = plot_scannable_positions(qcPath, geneName, scannableSequence, fraction_scannable)
		if plot_png:
			print("Scannability plot written: " + plot_png)
	except Exception as exc:   # a plotting hiccup must never fail the scan
		print("Note: could not render the scannability plot (%s)." % exc)

	return {"geneName": geneName, "output_dir": outputPath, "qc_dir": qcPath,
	        "plot_png": plot_png, "fraction_scannable": fraction_scannable, "summary": summary}


